import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import time
import os
import math



import numpy as np
import pyvisa
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter





# ============================================================
# BK PRECISION 895
# ============================================================



RESOURCE = "USB0::0x0471::0x2827::480K23104::INSTR"





class BK895App:



    def __init__(self, root):
        self.root = root
        self.root.title("BK Precision 895 - LCR Characterization")
        self.root.geometry("720x700")
        self.root.resizable(False, False)



        self.rm = None
        self.instrument = None
        self.stop_requested = False
        self.measuring = False



        self.start_freq = tk.StringVar(value="20")
        self.stop_freq = tk.StringVar(value="1M")
        self.points = tk.StringVar(value="21")
        self.scale = tk.StringVar(value="Logarithmic")
        self.voltage = tk.StringVar(value="0.005")
        self.cycles = tk.StringVar(value="2")
        self.delay = tk.StringVar(value="15")
        self.sample = tk.StringVar(value="sample")
        self.folder = tk.StringVar(value=os.path.abspath("."))
        self.cprp = tk.BooleanVar(value=True)
        self.csrs = tk.BooleanVar(value=True)
        self.status = tk.StringVar(value="Disconnected")
        self.progress_value = tk.DoubleVar(value=0)



        self.create_interface()



    # ========================================================
    # INTERFACE
    # ========================================================



    def create_interface(self):
        main = ttk.Frame(self.root, padding=15)
        main.pack(fill="both", expand=True)



        ttk.Label(
            main,
            text="BK Precision 895",
            font=("Segoe UI", 18, "bold")
        ).pack()



        ttk.Label(main, text="LCR Characterization").pack(pady=(0, 15))



        connection = ttk.LabelFrame(main, text="Instrument", padding=10)
        connection.pack(fill="x", pady=5)



        ttk.Label(connection, text="VISA Resource:").grid(row=0, column=0, sticky="w")



        self.resource_entry = ttk.Entry(connection, width=48)
        self.resource_entry.grid(row=0, column=1, padx=10)
        self.resource_entry.insert(0, RESOURCE)
        self.resource_entry.config(state="readonly")



        self.connect_button = ttk.Button(
            connection,
            text="CONNECT",
            command=self.connect
        )
        self.connect_button.grid(row=0, column=2)



        self.idn_label = ttk.Label(connection, text="Not connected")
        self.idn_label.grid(row=1, column=0, columnspan=3, sticky="w", pady=8)



        settings = ttk.LabelFrame(main, text="Frequency settings", padding=10)
        settings.pack(fill="x", pady=5)



        ttk.Label(settings, text="Start frequency:").grid(row=0, column=0, sticky="w")
        ttk.Entry(settings, textvariable=self.start_freq, width=15).grid(
            row=0, column=1, sticky="w", padx=10
        )
        ttk.Label(settings, text="Hz").grid(row=0, column=2)



        ttk.Label(settings, text="Stop frequency:").grid(
            row=0, column=3, sticky="w", padx=(30, 0)
        )
        ttk.Entry(settings, textvariable=self.stop_freq, width=15).grid(
            row=0, column=4, sticky="w", padx=10
        )
        ttk.Label(settings, text="Hz").grid(row=0, column=5)



        ttk.Label(settings, text="Number of points:").grid(
            row=1, column=0, sticky="w", pady=8
        )
        ttk.Entry(settings, textvariable=self.points, width=15).grid(
            row=1, column=1, sticky="w", padx=10
        )



        ttk.Label(settings, text="Scale:").grid(
            row=1, column=3, sticky="w", padx=(30, 0)
        )
        ttk.Combobox(
            settings,
            textvariable=self.scale,
            values=["Logarithmic", "Linear"],
            state="readonly",
            width=13
        ).grid(row=1, column=4, padx=10)



        measurement = ttk.LabelFrame(main, text="Measurement", padding=10)
        measurement.pack(fill="x", pady=5)



        ttk.Label(measurement, text="Voltage:").grid(row=0, column=0, sticky="w")
        ttk.Entry(measurement, textvariable=self.voltage, width=15).grid(
            row=0, column=1, padx=10
        )
        ttk.Label(measurement, text="V").grid(row=0, column=2)



        ttk.Label(measurement, text="Cycles:").grid(row=0, column=3, padx=(30, 0))
        ttk.Entry(measurement, textvariable=self.cycles, width=15).grid(
            row=0, column=4, padx=10
        )



        ttk.Label(measurement, text="Wait:").grid(row=1, column=0, sticky="w", pady=8)
        ttk.Entry(measurement, textvariable=self.delay, width=15).grid(
            row=1, column=1, padx=10
        )
        ttk.Label(measurement, text="seconds / frequency").grid(
            row=1, column=2, columnspan=3, sticky="w"
        )



        modes = ttk.LabelFrame(main, text="Measurement modes", padding=10)
        modes.pack(fill="x", pady=5)



        ttk.Checkbutton(
            modes,
            text="Parallel CP / RP",
            variable=self.cprp
        ).pack(side="left", padx=20)



        ttk.Checkbutton(
            modes,
            text="Series CS / RS",
            variable=self.csrs
        ).pack(side="left", padx=20)



        output = ttk.LabelFrame(main, text="Output", padding=10)
        output.pack(fill="x", pady=5)



        ttk.Label(output, text="Sample name:").grid(row=0, column=0, sticky="w")
        ttk.Entry(output, textvariable=self.sample, width=30).grid(
            row=0, column=1, padx=10
        )



        ttk.Label(output, text="Folder:").grid(row=1, column=0, sticky="w", pady=8)
        ttk.Entry(output, textvariable=self.folder, width=50).grid(
            row=1, column=1, padx=10
        )
        ttk.Button(output, text="Browse", command=self.choose_folder).grid(
            row=1, column=2
        )



        buttons = ttk.Frame(main)
        buttons.pack(pady=12)



        self.start_button = ttk.Button(
            buttons,
            text="START MEASUREMENT",
            command=self.start
        )
        self.start_button.pack(side="left", padx=5)



        self.stop_button = ttk.Button(
            buttons,
            text="STOP",
            command=self.stop,
            state="disabled"
        )
        self.stop_button.pack(side="left", padx=5)



        ttk.Label(main, textvariable=self.status).pack(anchor="w")



        ttk.Progressbar(
            main,
            variable=self.progress_value,
            maximum=100
        ).pack(fill="x", pady=5)



        log_frame = ttk.LabelFrame(main, text="Log", padding=5)
        log_frame.pack(fill="both", expand=True)



        self.log = tk.Text(log_frame, height=10, state="disabled")
        self.log.pack(fill="both", expand=True)



    # ========================================================
    # GUI HELPERS
    # ========================================================



    def write_log(self, text):
        def update():
            self.log.config(state="normal")
            self.log.insert("end", str(text) + "\n")
            self.log.see("end")
            self.log.config(state="disabled")



        self.root.after(0, update)



    def set_status(self, text):
        self.root.after(0, lambda: self.status.set(text))



    def set_progress(self, value):
        self.root.after(0, lambda: self.progress_value.set(value))



    # ========================================================
    # BK QUERY
    # ========================================================



    def query_bk(self, command):
        self.instrument.write(command)
        raw = self.instrument.read_raw()



        # The BK895 uses non-ASCII bytes for some symbols such as ohm.
        # We only need the numeric value and SI prefix, so those bytes
        # can safely be ignored here.
        return raw.decode("ascii", errors="ignore").strip()



    # ========================================================
    # CONNECT
    # ========================================================



    def connect(self):
        try:
            if self.instrument is not None:
                try:
                    self.instrument.close()
                except Exception:
                    pass
                self.instrument = None



            if self.rm is not None:
                try:
                    self.rm.close()
                except Exception:
                    pass
                self.rm = None



            self.rm = pyvisa.ResourceManager()
            resources = self.rm.list_resources()



            self.write_log("VISA instruments found:")
            for resource in resources:
                self.write_log("  " + resource)



            self.instrument = self.rm.open_resource(RESOURCE)
            self.instrument.timeout = 30000
            self.instrument.write_termination = "\n"



            idn = self.query_bk("*IDN?")



            self.idn_label.config(text=idn)
            self.status.set("Connected")
            self.write_log("")
            self.write_log("Connected:")
            self.write_log(idn)
            self.connect_button.config(text="RECONNECT")



        except Exception as e:
            self.instrument = None
            self.status.set("Connection error")
            self.write_log("Connection ERROR: " + str(e))
            messagebox.showerror("Connection error", str(e))



    # ========================================================
    # FOLDER
    # ========================================================



    def choose_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder.set(folder)



    # ========================================================
    # PARSERS
    # ========================================================



    def parse_frequency(self, value):
        value = value.strip().lower().replace(" ", "")



        if value.endswith("mhz"):
            return float(value[:-3]) * 1e6
        if value.endswith("khz"):
            return float(value[:-3]) * 1e3
        if value.endswith("hz"):
            return float(value[:-2])
        if value.endswith("m"):
            return float(value[:-1]) * 1e6
        if value.endswith("k"):
            return float(value[:-1]) * 1e3



        return float(value)



    def parse_capacitance(self, value):
        value = value.strip()



        if "---" in value or value == "":
            return math.nan



        units = [
            ("pF", 1e-12),
            ("nF", 1e-9),
            ("uF", 1e-6),
            ("mF", 1e-3),
            ("F", 1.0),
        ]



        for unit, multiplier in units:
            if value.endswith(unit):
                number = value[:-len(unit)].strip()
                return float(number) * multiplier



        raise ValueError("Unknown capacitance value: " + repr(value))



    def parse_resistance(self, value):
        value = value.strip()



        if "---" in value or value == "":
            return math.nan



        value = value.replace("Ohm", "").replace("OHM", "").strip()



        if value.endswith("G"):
            return float(value[:-1]) * 1e9
        if value.endswith("M"):
            return float(value[:-1]) * 1e6
        if value.endswith("k") or value.endswith("K"):
            return float(value[:-1]) * 1e3
        if value.endswith("m"):
            return float(value[:-1]) * 1e-3



        return float(value)



    def fetch_measurement(self):
        response = self.query_bk("FETC?")
        values = response.split(",")



        if len(values) < 2:
            raise RuntimeError("Unexpected FETC response: " + repr(response))



        capacitance_text = values[0].strip()
        resistance_text = values[1].strip()
        status_text = values[2].strip() if len(values) >= 3 else ""



        capacitance = self.parse_capacitance(capacitance_text)
        resistance = self.parse_resistance(resistance_text)



        return capacitance, resistance, status_text, response



    # ========================================================
    # FREQUENCIES
    # ========================================================



    def generate_frequencies(self):
        start = self.parse_frequency(self.start_freq.get())
        stop = self.parse_frequency(self.stop_freq.get())
        points = int(self.points.get())



        if start <= 0:
            raise ValueError("Start frequency must be > 0")
        if stop <= start:
            raise ValueError("Stop frequency must be greater than start frequency")
        if points < 2:
            raise ValueError("Number of points must be at least 2")



        if self.scale.get() == "Logarithmic":
            return np.geomspace(start, stop, points).tolist()



        return np.linspace(start, stop, points).tolist()



    # ========================================================
    # START / STOP
    # ========================================================



    def start(self):
        if self.instrument is None:
            messagebox.showwarning(
                "Not connected",
                "Connect the BK Precision 895 first."
            )
            return



        try:
            frequencies = self.generate_frequencies()
            voltage = float(self.voltage.get())
            cycles = int(self.cycles.get())
            delay = float(self.delay.get())
            sample = self.sample.get().strip()
            folder_base = self.folder.get()
            use_cprp = self.cprp.get()
            use_csrs = self.csrs.get()
            logarithmic_scale = self.scale.get() == "Logarithmic"



            if voltage <= 0:
                raise ValueError("Voltage must be > 0.")
            if cycles < 1:
                raise ValueError("Cycles must be at least 1.")
            if delay < 0:
                raise ValueError("Wait time cannot be negative.")
            if not use_cprp and not use_csrs:
                raise ValueError("Select at least one measurement mode.")
            if sample == "":
                raise ValueError("Enter a sample name.")



        except Exception as e:
            messagebox.showerror("Invalid settings", str(e))
            return



        self.stop_requested = False
        self.measuring = True
        self.progress_value.set(0)
        self.start_button.config(state="disabled")
        self.stop_button.config(state="normal")



        thread = threading.Thread(
            target=self.measurement,
            args=(
                frequencies,
                voltage,
                cycles,
                delay,
                sample,
                folder_base,
                use_cprp,
                use_csrs,
                logarithmic_scale,
            ),
            daemon=True,
        )
        thread.start()



    def stop(self):
        self.stop_requested = True
        self.status.set("Stopping...")
        self.write_log("STOP requested.")



    def wait(self, seconds):
        end = time.time() + seconds



        while time.time() < end:
            if self.stop_requested:
                return False
            time.sleep(0.1)



        return True



    # ========================================================
    # MEASURE ONE MODE
    # ========================================================



    def measure_mode(
        self,
        mode,
        mode_name,
        file_prefix,
        frequencies,
        cycles,
        delay,
        folder,
        total,
        completed,
    ):
        self.instrument.write("*CLS")
        self.instrument.write("FUNC:IMP " + mode)



        self.write_log("")
        self.write_log("===== " + mode_name + " =====")



        all_cycles = []



        for cycle in range(1, cycles + 1):
            if self.stop_requested:
                break



            cycle_data = []



            filename = os.path.join(
                folder,
                "%s_%d.txt" % (file_prefix, cycle)
            )



            with open(filename, "w", encoding="utf-8") as file:
                file.write("# BK Precision 895\n")
                file.write("# Mode: %s\n" % mode)
                file.write(
                    "# Frequency_Hz Capacitance_F Resistance_Ohm Status\n"
                )



                for freq in frequencies:
                    if self.stop_requested:
                        break



                    self.instrument.write("FREQ %.9e" % freq)



                    if not self.wait(delay):
                        break



                    capacitance, resistance, status_code, raw_response = (
                        self.fetch_measurement()
                    )



                    file.write(
                        "%.9e %.9e %.9e %s\n"
                        % (freq, capacitance, resistance, status_code)
                    )
                    file.flush()



                    cycle_data.append(
                        {
                            "frequency": freq,
                            "capacitance": capacitance,
                            "resistance": resistance,
                            "status": status_code,
                        }
                    )



                    completed += 1
                    percent = completed / total * 100
                    self.set_progress(percent)
                    self.set_status("%s - %.3f Hz" % (mode, freq))



                    c_display = (
                        "NaN" if math.isnan(capacitance)
                        else "%.6e F" % capacitance
                    )
                    r_display = (
                        "NaN" if math.isnan(resistance)
                        else "%.6e Ohm" % resistance
                    )



                    self.write_log(
                        "%s | Cycle %d | %.3f Hz | C=%s | R=%s | status=%s"
                        % (
                            mode,
                            cycle,
                            freq,
                            c_display,
                            r_display,
                            status_code,
                        )
                    )
                    self.write_log("    BK: " + raw_response)



            all_cycles.append(cycle_data)
            self.write_log("Saved: " + filename)



        return completed, all_cycles



    # ========================================================
    # EXCEL CREATION
    # ========================================================



    def excel_value(self, value):
        """Excel cells should be empty for invalid NaN measurements."""
        if isinstance(value, float) and math.isnan(value):
            return None
        return value



    def create_excel(self, folder, sample, results, logarithmic_scale):
        excel_path = os.path.join(folder, sample + "_results.xlsx")



        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)



        for mode, cycles_data in results.items():
            if not cycles_data:
                continue



            ws = wb.create_sheet(title=mode)



            # ------------------------------------------------
            # HEADERS
            # ------------------------------------------------



            ws.cell(row=1, column=1, value="Frequency_Hz")



            for cycle_index in range(len(cycles_data)):
                c_col = 2 + cycle_index * 2
                r_col = c_col + 1



                ws.cell(
                    row=1,
                    column=c_col,
                    value=f"Capacitance_Cycle_{cycle_index + 1}_F",
                )
                ws.cell(
                    row=1,
                    column=r_col,
                    value=f"Resistance_Cycle_{cycle_index + 1}_Ohm",
                )



            for cell in ws[1]:
                cell.font = Font(bold=True)



            # ------------------------------------------------
            # DATA
            # ------------------------------------------------



            max_points = max(len(cycle) for cycle in cycles_data)



            for point_index in range(max_points):
                excel_row = point_index + 2



                # Frequency from first cycle that has this point.
                frequency = None
                for cycle_data in cycles_data:
                    if point_index < len(cycle_data):
                        frequency = cycle_data[point_index]["frequency"]
                        break



                if frequency is not None:
                    ws.cell(row=excel_row, column=1, value=frequency)



                for cycle_index, cycle_data in enumerate(cycles_data):
                    if point_index >= len(cycle_data):
                        continue



                    c_col = 2 + cycle_index * 2
                    r_col = c_col + 1
                    point = cycle_data[point_index]



                    ws.cell(
                        row=excel_row,
                        column=c_col,
                        value=self.excel_value(point["capacitance"]),
                    )
                    ws.cell(
                        row=excel_row,
                        column=r_col,
                        value=self.excel_value(point["resistance"]),
                    )



            # Scientific notation in data cells.
            for row in ws.iter_rows(min_row=2, max_row=max_points + 1):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = "0.000000E+00"



            # Useful widths.
            for col in range(1, 2 + len(cycles_data) * 2):
                ws.column_dimensions[get_column_letter(col)].width = 28



            # ------------------------------------------------
            # CAPACITANCE XY CHART
            # ------------------------------------------------



            chart_c = ScatterChart()
            chart_c.title = f"{mode} - Capacitance vs Frequency"
            chart_c.x_axis.title = "Frequency (Hz)"
            chart_c.y_axis.title = "Capacitance (F)"
            chart_c.height = 10
            chart_c.width = 19
            chart_c.legend.position = "r"



            if logarithmic_scale:
                chart_c.x_axis.scaling.logBase = 10



            xvalues = Reference(
                ws,
                min_col=1,
                min_row=2,
                max_row=max_points + 1,
            )



            for cycle_index in range(len(cycles_data)):
                c_col = 2 + cycle_index * 2
                yvalues = Reference(
                    ws,
                    min_col=c_col,
                    min_row=2,
                    max_row=max_points + 1,
                )



                series = Series(
                    yvalues,
                    xvalues,
                    title=f"Cycle {cycle_index + 1}",
                )
                series.marker.symbol = "circle"
                series.graphicalProperties.line.noFill = False
                chart_c.series.append(series)



            ws.add_chart(chart_c, "A16")



            # ------------------------------------------------
            # RESISTANCE XY CHART
            # ------------------------------------------------



            chart_r = ScatterChart()
            chart_r.title = f"{mode} - Resistance vs Frequency"
            chart_r.x_axis.title = "Frequency (Hz)"
            chart_r.y_axis.title = "Resistance (Ohm)"
            chart_r.height = 10
            chart_r.width = 19
            chart_r.legend.position = "r"



            if logarithmic_scale:
                chart_r.x_axis.scaling.logBase = 10



            for cycle_index in range(len(cycles_data)):
                r_col = 3 + cycle_index * 2
                yvalues = Reference(
                    ws,
                    min_col=r_col,
                    min_row=2,
                    max_row=max_points + 1,
                )



                series = Series(
                    yvalues,
                    xvalues,
                    title=f"Cycle {cycle_index + 1}",
                )
                series.marker.symbol = "circle"
                series.graphicalProperties.line.noFill = False
                chart_r.series.append(series)



            ws.add_chart(chart_r, "J16")



            ws.freeze_panes = "A2"



        if not wb.sheetnames:
            ws = wb.create_sheet("Results")
            ws["A1"] = "No measurement data available."



        wb.save(excel_path)
        return excel_path



    # ========================================================
    # COMPLETE MEASUREMENT
    # ========================================================



    def measurement(
        self,
        frequencies,
        voltage,
        cycles,
        delay,
        sample,
        folder_base,
        use_cprp,
        use_csrs,
        logarithmic_scale,
    ):
        try:
            folder = os.path.join(folder_base, sample)
            os.makedirs(folder, exist_ok=True)



            modes = int(use_cprp) + int(use_csrs)
            total = len(frequencies) * cycles * modes
            completed = 0



            results = {}



            self.instrument.write("*CLS")
            self.instrument.write("VOLTage %.9g" % voltage)
            self.instrument.write("APER FAST,1")



            self.write_log("")
            self.write_log("==============================")
            self.write_log("===== START MEASUREMENT =====")
            self.write_log("==============================")
            self.write_log("Sample: " + sample)
            self.write_log("Output: " + folder)
            self.write_log("Voltage: %.6g V" % voltage)
            self.write_log("Cycles: %d" % cycles)
            self.write_log("Frequencies: %d" % len(frequencies))
            self.write_log("Wait: %.3f s" % delay)



            if use_cprp and not self.stop_requested:
                completed, cprp_data = self.measure_mode(
                    mode="CPRP",
                    mode_name="CP / RP",
                    file_prefix="CP_RP_" + sample,
                    frequencies=frequencies,
                    cycles=cycles,
                    delay=delay,
                    folder=folder,
                    total=total,
                    completed=completed,
                )
                results["CPRP"] = cprp_data



            if use_csrs and not self.stop_requested:
                completed, csrs_data = self.measure_mode(
                    mode="CSRS",
                    mode_name="CS / RS",
                    file_prefix="CS_RS_" + sample,
                    frequencies=frequencies,
                    cycles=cycles,
                    delay=delay,
                    folder=folder,
                    total=total,
                    completed=completed,
                )
                results["CSRS"] = csrs_data



            self.instrument.write("*CLS")



            # Create Excel even if STOP was requested, using whatever
            # measurements were completed before stopping.
            if results:
                excel_path = self.create_excel(
                    folder,
                    sample,
                    results,
                    logarithmic_scale,
                )
                self.write_log("")
                self.write_log("Excel created:")
                self.write_log(excel_path)



            if self.stop_requested:
                self.set_status("Measurement stopped")
                self.write_log("")
                self.write_log("===== STOPPED =====")
            else:
                self.set_progress(100)
                self.set_status("Measurement completed")
                self.write_log("")
                self.write_log("===== COMPLETED =====")
                self.write_log("Results saved in:")
                self.write_log(folder)



        except Exception as e:
            self.set_status("Measurement error")
            self.write_log("")
            self.write_log("ERROR:")
            self.write_log(str(e))



            error_text = str(e)
            self.root.after(
                0,
                lambda text=error_text: messagebox.showerror(
                    "Measurement error",
                    text,
                ),
            )



        finally:
            self.measuring = False
            self.root.after(0, self.measurement_finished)



    # ========================================================
    # FINISH / CLOSE
    # ========================================================



    def measurement_finished(self):
        self.start_button.config(state="normal")
        self.stop_button.config(state="disabled")



    def close(self):
        if self.measuring:
            answer = messagebox.askyesno(
                "Measurement running",
                "A measurement is running. Stop and close?",
            )



            if not answer:
                return



            self.stop_requested = True



        try:
            if self.instrument:
                self.instrument.close()
            if self.rm:
                self.rm.close()
        except Exception:
            pass



        self.root.destroy()





# ============================================================
# MAIN
# ============================================================



if __name__ == "__main__":
    root = tk.Tk()
    app = BK895App(root)
    root.protocol("WM_DELETE_WINDOW", app.close)
    root.mainloop()
 